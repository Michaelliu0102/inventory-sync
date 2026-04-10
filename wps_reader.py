"""
WPS Docs 读取器 - 通过下载链接获取 xlsx 文件并解析中国库存数据
"""

import io
import logging
import tempfile

import requests
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def read_inventory(config: dict) -> dict:
    """
    从 WPS Docs 下载 xlsx 文件并读取库存数据。

    Returns:
        {"DisplayName1": 10, "DisplayName2": 5, ...}
    """
    wps = config["wps"]
    download_url = wps["download_url"]
    name_col = wps.get("name_column", 2)       # 默认 B 列
    qty_col = wps.get("quantity_column", 3)     # 默认 C 列
    header_rows = wps.get("header_rows", 1)

    logger.info(f"Loading WPS Excel file from: {download_url}")
    
    if download_url.startswith("http"):
        resp = requests.get(download_url, timeout=60)
        resp.raise_for_status()
        
        # 很多时候分享链接下载下来的是网页而不是真实的 xlsx 文件，加个防护提示
        if b"<html" in resp.content[:1000].lower():
            logger.error("The downloaded content is an HTML webpage, not an Excel file.")
            logger.error("Please provide a *DIRECT* download link, or just download the file manually and put the local file path (e.g. 'china.xlsx') in config.yaml.")
            raise ValueError("WPS link returned an HTML page instead of an Excel file.")
            
        file_bytes = io.BytesIO(resp.content)
    else:
        # 如果填的不是 http 打头的网址，就当成本地文件路径处理
        with open(download_url, "rb") as f:
            file_bytes = io.BytesIO(f.read())

    # 加载 xlsx
    wb = load_workbook(filename=file_bytes, read_only=True, data_only=True)
    ws = wb.active

    logger.info(f"Reading WPS sheet: {ws.title}")

    inventory = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
        if i <= header_rows:
            continue

        # 列号从 1 开始
        if len(row) < max(name_col, qty_col):
            continue

        name = row[name_col - 1]
        qty_raw = row[qty_col - 1]

        if not name:
            continue

        name = str(name).strip()

        try:
            quantity = float(str(qty_raw).replace(",", "").strip()) if qty_raw is not None else 0
        except (ValueError, TypeError):
            logger.warning(f"  Row {i}: cannot parse quantity '{qty_raw}' for '{name}', skipping")
            continue

        inventory[name] = inventory.get(name, 0) + quantity

    wb.close()
    logger.info(f"  WPS Docs: {len(inventory)} SKUs loaded")
    return inventory
